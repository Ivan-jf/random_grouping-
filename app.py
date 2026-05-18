import os
import uuid
import random
import re
import pandas as pd
from flask import Flask, render_template, request, send_file, jsonify
from werkzeug.utils import secure_filename
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter

app = Flask(__name__)
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
app.config['UPLOAD_FOLDER'] = os.path.join(BASE_DIR, 'uploads')
app.config['OUTPUT_FOLDER'] = os.path.join(BASE_DIR, 'outputs')
app.config['MAX_CONTENT_LENGTH'] = 16 * 1024 * 1024  # 16MB

# 自动创建必要目录
os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)
os.makedirs(app.config['OUTPUT_FOLDER'], exist_ok=True)

ALLOWED_EXTENSIONS = {'xlsx', 'xls'}


def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS


def clean_value(v):
    """将 pandas / numpy 中不适合 JSON 的值转成普通 Python 值"""
    if pd.isna(v):
        return None
    return v


def dataframe_to_records(df):
    """DataFrame 转 records，并处理 NaN"""
    records = []
    for _, row in df.iterrows():
        item = {}
        for col in df.columns:
            item[col] = clean_value(row[col])
        records.append(item)
    return records


def random_grouping(df, group_num, group_names, id_col, var_col, filter_col=None):
    """
    随机分组核心逻辑

    注意：
    - group_num: 每组人数
    - group_names: 分组名称列表
    - id_col: 样品 ID 列名
    - var_col: 用于排序分组的变量列名
    - filter_col: "要不要"列名，只取 Y 的行，可选
    """

    if not filter_col and '要不要' in df.columns:
        filter_col = '要不要'

    if id_col not in df.columns:
        raise ValueError(f"ID 列不存在：{id_col}")

    if var_col not in df.columns:
        raise ValueError(f"排序变量列不存在：{var_col}")

    if filter_col and filter_col not in df.columns:
        raise ValueError(f"筛选列不存在：{filter_col}")

    # 过滤"要不要"列
    if filter_col and filter_col in df.columns:
        participate_mask = df[filter_col].astype(str).str.strip().str.upper() == 'Y'
        df_excluded = df[~participate_mask].copy()
        df_excluded['未参与原因'] = f'{filter_col} 列不是 Y'
        df = df[participate_mask].copy()
        df = df.reset_index(drop=True)
    else:
        df_excluded = df.iloc[0:0].copy()
        df_excluded['未参与原因'] = '未选择筛选列'

    n = len(df)
    group_count = len(group_names)
    expected = group_num * group_count

    duplicated_id_values = df[df[id_col].duplicated(keep=False)][id_col]
    if not duplicated_id_values.empty:
        duplicate_preview = duplicated_id_values.drop_duplicates().head(10).tolist()
        duplicate_text = '、'.join(str(v) for v in duplicate_preview)
        more_text = ' 等' if duplicated_id_values.nunique(dropna=False) > len(duplicate_preview) else ''
        raise ValueError(
            f"样品 ID 列「{id_col}」存在重复值：{duplicate_text}{more_text}。"
            "请先处理重复 ID 后再分组，确保每个样品都有唯一标识。"
        )

    if n != expected:
        raise ValueError(
            f"筛选后有效行数为 {n}，但 每组人数({group_num}) × 组名数量({group_count}) = {expected}，"
            f"两者必须相等，请检查数据或参数。"
        )

    # 按 var_col 降序排列
    df_sorted = df.sort_values(by=var_col, ascending=False).reset_index(drop=True)

    # 分块：每个区块包含每个组各 1 只动物，所以每块大小 = 组别数量
    block_size = group_count
    df_sorted['block'] = [i // block_size + 1 for i in range(n)]

    # 每块内随机排序并分配组名
    df_sorted['group'] = None

    for block_id in df_sorted['block'].unique():
        idx = df_sorted[df_sorted['block'] == block_id].index.tolist()
        shuffled = idx.copy()
        random.shuffle(shuffled)

        # 每个 block 内，每个组名出现一次
        names_for_block = group_names[:len(idx)]

        for pos, original_idx in enumerate(shuffled):
            df_sorted.at[original_idx, 'group'] = names_for_block[pos]

    # 统计摘要
    summary_df = df_sorted.groupby('group')[var_col].agg(
        mean='mean',
        sd='std',
        min='min',
        max='max'
    ).reset_index()

    summary_df.columns = ['分组', '均值', '标准差', '最小值', '最大值']
    summary_df = summary_df.round(4)

    return df_sorted, summary_df, df_excluded


def move_export_columns(df, group_col='group', cage_col='笼号'):
    """将 group 放到第一列；如果存在笼号列，则放到 group 后面。"""
    cols = list(df.columns)

    if group_col not in cols:
        return df

    cols.remove(group_col)
    cols.insert(0, group_col)

    if cage_col in cols:
        cols.remove(cage_col)
        cols.insert(1, cage_col)

    return df[cols]


def apply_group_row_colors(ws, group_col_name='group', cage_col_name='笼号'):
    """根据 group 给整行填浅色；如果有笼号列，再按笼号给该单元格填深色。"""
    headers = [cell.value for cell in ws[1]]

    if group_col_name not in headers:
        return

    group_col_idx = headers.index(group_col_name) + 1
    cage_col_idx = headers.index(cage_col_name) + 1 if cage_col_name in headers else None

    colors = [
        "D9EAF7",  # 浅蓝
        "DFF3E3",  # 浅绿
        "FFF2CC",  # 浅黄
        "F4DDEB",  # 浅粉
        "E8DAEF",  # 浅紫
        "D6EAF8",  # 天蓝
        "FADBD8",  # 浅红
        "D5F5E3",  # 薄荷绿
        "FCF3CF",  # 米黄
        "EAECEE",  # 浅灰
        "D1F2EB",  # 青绿
        "FDEBD0",  # 浅橙
    ]

    cage_colors = [
        "1F4E79",  # 深蓝
        "375623",  # 深绿
        "7F6000",  # 深金
        "7030A0",  # 深紫
        "833C0C",  # 深棕
        "C00000",  # 深红
        "0F6B78",  # 深青
        "404040",  # 深灰
        "5B2C6F",  # 暗紫
        "145A32",  # 暗绿
    ]

    group_fill_map = {}
    cage_fill_map = {}
    color_index = 0
    cage_color_index = 0

    # 设置表头样式
    header_fill = PatternFill(fill_type="solid", fgColor="2B6CB0")
    header_font = Font(color="FFFFFF", bold=True)
    header_alignment = Alignment(horizontal="center", vertical="center")

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment

    cage_font = Font(color="FFFFFF", bold=True)

    # 给每个组别分配浅色；笼号列按笼号值额外分配深色
    for row_idx in range(2, ws.max_row + 1):
        group_value = ws.cell(row=row_idx, column=group_col_idx).value

        if group_value not in group_fill_map:
            color = colors[color_index % len(colors)]
            group_fill_map[group_value] = PatternFill(
                fill_type="solid",
                fgColor=color
            )
            color_index += 1

        fill = group_fill_map[group_value]

        for col_idx in range(1, ws.max_column + 1):
            ws.cell(row=row_idx, column=col_idx).fill = fill
            ws.cell(row=row_idx, column=col_idx).alignment = Alignment(
                vertical="center",
                wrap_text=True
            )

        if cage_col_idx:
            cage_cell = ws.cell(row=row_idx, column=cage_col_idx)
            cage_value = cage_cell.value

            if cage_value is not None and str(cage_value).strip() != "":
                if cage_value not in cage_fill_map:
                    cage_fill_map[cage_value] = PatternFill(
                        fill_type="solid",
                        fgColor=cage_colors[cage_color_index % len(cage_colors)]
                    )
                    cage_color_index += 1

                cage_cell.fill = cage_fill_map[cage_value]
                cage_cell.font = cage_font
                cage_cell.alignment = Alignment(
                    horizontal="center",
                    vertical="center",
                    wrap_text=True
                )

    # 冻结首行
    ws.freeze_panes = "A2"

    # 添加筛选
    ws.auto_filter.ref = ws.dimensions

    # 自动调整列宽
    for col_idx in range(1, ws.max_column + 1):
        col_letter = get_column_letter(col_idx)
        max_length = 0

        for cell in ws[col_letter]:
            value = cell.value
            if value is None:
                continue

            value_length = len(str(value))
            if value_length > max_length:
                max_length = value_length

        ws.column_dimensions[col_letter].width = min(max_length + 4, 35)


@app.route('/')
def index():
    return render_template('index.html')


@app.route('/get_columns', methods=['POST'])
def get_columns():
    """上传文件后返回列名供用户选择"""
    if 'file' not in request.files:
        return jsonify({'error': '未上传文件'}), 400

    file = request.files['file']

    if file.filename == '':
        return jsonify({'error': '文件名为空'}), 400

    if not allowed_file(file.filename):
        return jsonify({'error': '仅支持 .xlsx / .xls 格式'}), 400

    filename = secure_filename(file.filename)
    uid = str(uuid.uuid4())[:8]
    save_path = os.path.join(app.config['UPLOAD_FOLDER'], uid + '_' + filename)
    file.save(save_path)

    try:
        df = pd.read_excel(save_path)
        columns = df.columns.tolist()
        return jsonify({'columns': columns, 'filepath': save_path})
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/run', methods=['POST'])
def run():
    data = request.get_json()

    filepath = data.get('filepath')
    try:
        group_num = int(data.get('group_num'))
    except (TypeError, ValueError):
        return jsonify({'error': '请输入有效的每组人数'}), 400

    group_names = [g.strip() for g in re.split(r'[,，]', data.get('group_names', '')) if g.strip()]
    id_col = data.get('id_col')
    var_col = data.get('var_col')
    filter_col = data.get('filter_col') or None

    if not filepath or not os.path.exists(filepath):
        return jsonify({'error': '找不到上传的文件，请重新上传'}), 400

    if not group_names:
        return jsonify({'error': '请输入分组名称'}), 400

    duplicated_group_names = sorted({name for name in group_names if group_names.count(name) > 1})
    if duplicated_group_names:
        return jsonify({'error': f"分组名称不能重复：{'、'.join(duplicated_group_names)}"}), 400

    if not id_col:
        return jsonify({'error': '请选择样品 ID 列'}), 400

    if not var_col:
        return jsonify({'error': '请选择排序变量列'}), 400

    if group_num < 1:
        return jsonify({'error': '每组人数必须大于等于 1'}), 400

    try:
        df = pd.read_excel(filepath)
        source_total = len(df)

        df_result, summary_df, df_excluded = random_grouping(
            df=df,
            group_num=group_num,
            group_names=group_names,
            id_col=id_col,
            var_col=var_col,
            filter_col=filter_col
        )

        # 将 group 列移动到最前面；如果有笼号列，则放到 group 后面
        df_result_export = move_export_columns(df_result, group_col='group', cage_col='笼号')

        # 保存结果
        out_name = 'result_' + str(uuid.uuid4())[:8] + '.xlsx'
        out_path = os.path.join(app.config['OUTPUT_FOLDER'], out_name)

        with pd.ExcelWriter(out_path, engine='openpyxl') as writer:
            df_result_export.to_excel(writer, sheet_name='分组结果', index=False)
            summary_df.to_excel(writer, sheet_name='统计摘要', index=False)

            df_excluded.to_excel(writer, sheet_name='未参与样本', index=False)

            # 给“分组结果”sheet 按组别上色
            ws = writer.sheets['分组结果']
            apply_group_row_colors(ws, group_col_name='group')

        # 返回前端预览数据
        summary_records = dataframe_to_records(summary_df)

        preview_cols = ['group', id_col, var_col, 'block']
        preview_cols = [c for c in preview_cols if c in df_result_export.columns]
        preview = dataframe_to_records(df_result_export[preview_cols].head(20))

        return jsonify({
            'success': True,
            'out_file': out_name,
            'summary': summary_records,
            'preview': preview,
            'source_total': source_total,
            'total': len(df_result_export),
            'excluded_total': len(df_excluded)
        })

    except ValueError as e:
        return jsonify({'error': str(e)}), 400
    except Exception as e:
        return jsonify({'error': f'运行出错：{str(e)}'}), 500


@app.route('/download/<filename>')
def download(filename):
    filename = secure_filename(filename)
    path = os.path.join(app.config['OUTPUT_FOLDER'], filename)

    if not os.path.exists(path):
        return '文件不存在', 404

    return send_file(path, as_attachment=True, download_name=filename)


if __name__ == '__main__':
    app.run(debug=True)
